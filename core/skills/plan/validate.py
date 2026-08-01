#!/usr/bin/env python3
"""Gate 9 validator for a plan workbook. Run until clean before delivering.

Catches the failures that are invisible on screen but wrong in substance:

  * dangling / self / cyclic dependencies
  * tasks over the size cap
  * rows missing priority, phase, owner, estimate or acceptance criteria
  * duplicate IDs
  * ORPHANED REFERENCES — another sheet citing a task that was deferred or deleted.
    Deferring work silently falsifies risks whose mitigation WAS that task, and
    checklist lines that ask for evidence nothing will now produce.
  * stale terminology left behind after a rename (Wave -> Phase)
  * fractional days rendered with a 0-decimal number format (0.5 showing as "0")

    validate.py --backlog plan.xlsx
    validate.py --backlog plan.xlsx --deferred-sheet "02b Deferred Improvements" \\
                --stale Wave --stale "S/M/L/XL" --cap 5
"""

import argparse
import re
import sys
from collections import defaultdict

ID_RE = re.compile(r"\b[A-Z]{1,2}-\d{1,3}[a-z]?\b")
ALIASES = {
    "id": ["id", "task id", "ref"],
    "task": ["task", "title", "name", "description"],
    "days": ["effort (days)", "days", "effort", "estimate (days)", "man days"],
    "phase": ["phase", "wave", "sprint", "release"],
    "role": ["role", "owner", "assignee", "seat"],
    "deps": ["depends on", "dependencies", "dependency", "predecessor"],
    "prio": ["priority", "prio"],
    "accept": ["acceptance criteria", "acceptance", "definition of done"],
}
FAIL, WARN = [], []


def fail(msg):
    FAIL.append(msg)


def warn(msg):
    WARN.append(msg)


def find_backlog(wb, name):
    if name in wb.sheetnames:
        return name
    c = [s for s in wb.sheetnames if "backlog" in s.lower()]
    if len(c) == 1:
        return c[0]
    sys.exit(f"error: cannot find the backlog sheet. Sheets: {wb.sheetnames}")


def main():
    import openpyxl

    ap = argparse.ArgumentParser(
        description=__doc__, formatter_class=argparse.RawDescriptionHelpFormatter
    )
    ap.add_argument("--backlog", required=True)
    ap.add_argument("--sheet", default="02 Task Backlog")
    ap.add_argument("--deferred-sheet", default="02b Deferred Improvements")
    ap.add_argument("--cap", type=float, default=5.0)
    ap.add_argument("--stale", action="append", default=[])
    a = ap.parse_args()

    wb = openpyxl.load_workbook(a.backlog)
    sheet = find_backlog(wb, a.sheet)
    ws = wb[sheet]

    hdr = None
    for r in range(1, min(ws.max_row, 12) + 1):
        low = [
            str(ws.cell(row=r, column=c).value).strip().lower()
            for c in range(1, ws.max_column + 1)
            if ws.cell(row=r, column=c).value is not None
        ]
        if any(v in ALIASES["id"] for v in low) and any(
            v in ALIASES["days"] for v in low
        ):
            hdr = r
            break
    if hdr is None:
        sys.exit("error: no header row found")
    heads = {
        str(ws.cell(row=hdr, column=c).value).strip().lower(): c
        for c in range(1, ws.max_column + 1)
        if ws.cell(row=hdr, column=c).value is not None
    }
    M = {}
    for k, names in ALIASES.items():
        for n in names:
            if n in heads:
                M[k] = heads[n]
                break
    missing_cols = [k for k in ("id", "days") if k not in M]
    if missing_cols:
        sys.exit(f"error: backlog has no {missing_cols} column")

    rows, seen = {}, defaultdict(list)
    for r in range(hdr + 1, ws.max_row + 1):
        tid = ws.cell(row=r, column=M["id"]).value
        if not tid:
            continue
        tid = str(tid).strip()
        seen[tid].append(r)
        cell = lambda k: ws.cell(row=r, column=M[k]).value if k in M else None  # noqa: E731
        rows[tid] = {"row": r, **{k: cell(k) for k in M}}
    for tid, rs in seen.items():
        if len(rs) > 1:
            fail(f"duplicate ID {tid} on rows {rs}")

    active = {t: v for t, v in rows.items() if isinstance(v["days"], (int, float))}
    informational = set(rows) - set(active)

    # ---- size cap, required fields ----
    for tid, v in active.items():
        if v["days"] > a.cap:
            fail(
                f"{tid} is {v['days']:g} days — over the {a.cap:g}-day cap. Split it along a real seam."
            )
        if v["days"] <= 0:
            fail(f"{tid} has an effort of {v['days']}")
        for k, label in (
            ("prio", "priority"),
            ("phase", "phase"),
            ("role", "owner"),
            ("accept", "acceptance criteria"),
        ):
            if k in M and (v.get(k) in (None, "", "—")):
                warn(f"{tid} has no {label}")
        if "days" in M:
            nf = ws.cell(row=v["row"], column=M["days"]).number_format
            if float(v["days"]) % 1 and nf and re.fullmatch(r"[#0,]+", nf):
                fail(
                    f"{tid} is {v['days']:g} days but its number format {nf!r} hides the decimal"
                )

    # ---- dependency graph ----
    kids = defaultdict(list)
    for tid, v in active.items():
        raw = str(v.get("deps") or "")
        for d in ID_RE.findall(raw):
            if d == tid:
                fail(f"{tid} depends on itself")
            elif d in active:
                kids[d].append(tid)
            elif d in informational:
                fail(
                    f"{tid} depends on {d}, which carries no effort and will never complete"
                )
            elif re.match(r"^[A-Z]{2}-", d):
                pass  # decision reference, e.g. OD-11
            else:
                fail(f"{tid} depends on {d}, which is not in the backlog")
    colour, stack = defaultdict(int), []

    def visit(n):
        colour[n] = 1
        stack.append(n)
        for k in kids[n]:
            if colour[k] == 1:
                fail("dependency cycle: " + " -> ".join(stack[stack.index(k) :] + [k]))
            elif colour[k] == 0:
                visit(k)
        stack.pop()
        colour[n] = 2

    for n in active:
        if colour[n] == 0:
            visit(n)

    # ---- orphaned references to deferred / non-existent tasks ----
    deferred = set()
    if a.deferred_sheet in wb.sheetnames:
        d = wb[a.deferred_sheet]
        for r in range(1, d.max_row + 1):
            v = d.cell(row=r, column=1).value
            if v and ID_RE.fullmatch(str(v).strip()):
                deferred.add(str(v).strip())
    known = set(rows) | deferred
    # Only IDs whose PREFIX actually appears in the backlog are task references. Risk (R-),
    # decision (OD-) and ADR ids share the same shape and must not be reported as missing
    # tasks — derive the valid prefixes from the data rather than hardcoding exclusions.
    PREFIXES = {t.split("-")[0] for t in known}
    ALLOW = {a.deferred_sheet, sheet}
    for name in wb.sheetnames:
        if name in ALLOW:
            continue
        sh = wb[name]
        for r in range(1, sh.max_row + 1):
            for c in range(1, sh.max_column + 1):
                val = sh.cell(row=r, column=c).value
                if not isinstance(val, str):
                    continue
                for d in ID_RE.findall(val):
                    if d.split("-")[0] not in PREFIXES:
                        continue
                    if d in deferred:
                        warn(
                            f"{name}!{sh.cell(row=r, column=c).coordinate} cites {d}, which is "
                            f"DEFERRED — if this is a mitigation or a piece of evidence it is now false"
                        )
                    elif d not in known:
                        fail(
                            f"{name}!{sh.cell(row=r, column=c).coordinate} cites {d}, which "
                            f"does not exist anywhere"
                        )

    # ---- stale terminology ----
    for term in a.stale:
        hits = []
        for name in wb.sheetnames:
            sh = wb[name]
            for r in range(1, sh.max_row + 1):
                for c in range(1, sh.max_column + 1):
                    v = sh.cell(row=r, column=c).value
                    if isinstance(v, str) and term.lower() in v.lower():
                        hits.append(f"{name}!{sh.cell(row=r, column=c).coordinate}")
        if hits:
            fail(
                f"stale term {term!r} still appears in {len(hits)} place(s): {', '.join(hits[:6])}"
            )

    # ---- report ----
    total = sum(v["days"] for v in active.values())
    by_phase = defaultdict(float)
    for v in active.values():
        by_phase[str(v.get("phase") or "—")] += v["days"]
    print(f"backlog {sheet!r}: {len(active)} tasks · {total:g} days")
    if informational:
        print(
            f"  ({len(informational)} row(s) with no effort, excluded from every total: {', '.join(sorted(informational))})"
        )
    for p in sorted(by_phase):
        print(f"  {p:12} {by_phase[p]:>7g} d")
    if deferred:
        print(f"deferred: {len(deferred)} task(s) — {', '.join(sorted(deferred))}")
    print()
    for m in FAIL:
        print("  FAIL  " + m)
    for m in WARN:
        print("  warn  " + m)
    print()
    if FAIL:
        print(
            f"NOT DELIVERABLE — {len(FAIL)} failure(s), {len(WARN)} warning(s). Fix the failures."
        )
        sys.exit(1)
    print(
        f"PASSED — 0 failures, {len(WARN)} warning(s). Totals above are the authoritative figures; every other sheet must match them."
    )


if __name__ == "__main__":
    main()

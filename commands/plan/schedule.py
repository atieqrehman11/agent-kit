#!/usr/bin/env python3
"""Resource-constrained scheduler for a task backlog. Generic across projects.

Effort is not duration. This computes the real makespan: each seat works one task at a
time at its allocation (a 50% seat needs 2 calendar days per task-day), and a task starts
when its dependencies are done AND its owner is free. Ready tasks are ordered by longest
downstream chain, with phase as a tie-break priority rather than a hard gate.

Reads the backlog by COLUMN HEADING, never by position — users reorder columns and
position-based access silently returns the wrong column.

    # what does the plan actually take?
    schedule.py --backlog plan.xlsx --alloc team.json

    # try allocation changes before recommending one
    schedule.py --backlog plan.xlsx --alloc team.json --scenarios

    # level work across the pooled seats and write Role back to the backlog
    schedule.py --backlog plan.xlsx --alloc team.json --level --apply

    # write the Gantt sheet
    schedule.py --backlog plan.xlsx --alloc team.json --gantt "04b Gantt" --apply

team.json:
    {
      "roles": {
        "Full-stack 1":       {"alloc": 1.0,  "pool": true},
        "Platform & DevOps":  {"alloc": 0.5,  "pool": true},
        "QA / Evaluation":    {"alloc": 0.5,  "pool": false},
        "Legal SME (client)": {"alloc": 0.25, "pool": false}
      },
      "restricted":      {"Frontend": ["Full-stack 1", "Full-stack 2"]},
      "gate_last_phase": true,
      "loe_days_per_week": {"PM / Solution Owner": 2.5}
    }

  pool:false            keeps that role's own tasks (skill does not transfer)
  restricted            workstream -> only these seats may take it
  gate_last_phase       the final phase waits for every earlier phase to finish
  loe_days_per_week     continuous roles budgeted outside the backlog, reported separately
"""

import argparse
import json
import re
import shutil
import sys
from collections import defaultdict
from functools import lru_cache
from pathlib import Path

ID_RE = re.compile(r"\b[A-Z]{1,2}-\d{1,3}[a-z]?\b")
ALIASES = {
    "id": ["id", "task id", "ref"],
    "task": ["task", "title", "name", "description"],
    "days": ["effort (days)", "days", "effort", "estimate (days)", "man days"],
    "phase": ["phase", "wave", "sprint", "release"],
    "role": ["role", "owner", "assignee", "seat"],
    "deps": ["depends on", "dependencies", "dependency", "predecessor"],
    "ws": ["workstream", "stream", "category", "area"],
    "prio": ["priority", "prio"],
    "accept": ["acceptance criteria", "acceptance", "definition of done"],
    "why": ["why it is needed for production", "why it is needed", "rationale", "why"],
}


def col_map(headings):
    """Map logical field -> 0-based column index, by heading text."""
    low = {str(h).strip().lower(): i for i, h in enumerate(headings) if h is not None}
    out = {}
    for key, names in ALIASES.items():
        for n in names:
            if n in low:
                out[key] = low[n]
                break
    for required in ("id", "days"):
        if required not in out:
            sys.exit(
                f"error: no column found for '{required}'. Looked for "
                f"{ALIASES[required]}. Found: {sorted(low)}"
            )
    return out


def load(path, sheet):
    import openpyxl

    wb = openpyxl.load_workbook(path)
    if sheet not in wb.sheetnames:
        cands = [s for s in wb.sheetnames if "backlog" in s.lower()]
        if len(cands) != 1:
            sys.exit(f"error: sheet {sheet!r} not found. Sheets: {wb.sheetnames}")
        sheet = cands[0]
    ws = wb[sheet]
    hdr_row = None
    for r in range(1, min(ws.max_row, 12) + 1):
        vals = [ws.cell(row=r, column=c).value for c in range(1, ws.max_column + 1)]
        low = [str(v).strip().lower() for v in vals if v is not None]
        if any(v in ALIASES["id"] for v in low) and any(
            v in ALIASES["days"] for v in low
        ):
            hdr_row = r
            break
    if hdr_row is None:
        sys.exit(
            "error: could not find a header row containing an ID and an effort column"
        )
    headings = [
        ws.cell(row=hdr_row, column=c).value for c in range(1, ws.max_column + 1)
    ]
    M = col_map(headings)
    tasks = {}
    for r in range(hdr_row + 1, ws.max_row + 1):
        row = [ws.cell(row=r, column=c).value for c in range(1, ws.max_column + 1)]
        tid = row[M["id"]]
        if not tid:
            continue
        d = row[M["days"]]
        if not isinstance(d, (int, float)):
            continue  # rejected / informational rows carry no effort
        get = lambda k, dflt="": (  # noqa: E731
            str(row[M[k]]) if k in M and row[M[k]] is not None else dflt
        )
        tasks[str(tid)] = {
            "id": str(tid),
            "task": get("task"),
            "days": float(d),
            "phase": get("phase", "—"),
            "role0": get("role", "—"),
            "ws": get("ws", "—"),
            "prio": get("prio", ""),
            "deps_raw": get("deps"),
            "accept": get("accept"),
            "why": get("why"),
        }
    return wb, sheet, ws, hdr_row, M, tasks


def wire(tasks):
    """Resolve dependency IDs and validate the graph. Exits on a broken graph."""
    problems = []
    for t in tasks.values():
        found = ID_RE.findall(t["deps_raw"])
        t["deps"] = [d for d in found if d in tasks and d != t["id"]]
        for d in found:
            # a 2-letter prefix is a decision reference (OD-11), not a task
            if d not in tasks and not re.match(r"^[A-Z]{2}-", d):
                problems.append(
                    f"{t['id']} depends on {d}, which is not in the backlog"
                )
    kids = defaultdict(list)
    for t in tasks.values():
        for d in t["deps"]:
            kids[d].append(t["id"])
    # cycle check
    WHITE, GREY, BLACK = 0, 1, 2
    colour = defaultdict(int)
    stack = []

    def visit(n):
        colour[n] = GREY
        stack.append(n)
        for k in kids[n]:
            if colour[k] == GREY:
                i = stack.index(k)
                problems.append("dependency cycle: " + " -> ".join(stack[i:] + [k]))
            elif colour[k] == WHITE:
                visit(k)
        stack.pop()
        colour[n] = BLACK

    for n in tasks:
        if colour[n] == WHITE:
            visit(n)
    return kids, problems


def phase_order(tasks):
    seen = sorted({t["phase"] for t in tasks.values()})
    named = [p for p in seen if re.search(r"\d", p)]
    named.sort(key=lambda p: [int(x) for x in re.findall(r"\d+", p)])
    rest = [p for p in seen if p not in named]
    order = {p: i + 1 for i, p in enumerate(named)}
    for p in rest:
        order[p] = 99
    return order, named


def run(tasks, kids, cfg, level=False):
    roles = cfg["roles"]
    alloc = {r: v["alloc"] for r, v in roles.items()}
    pool = [r for r, v in roles.items() if v.get("pool")]
    restricted = cfg.get("restricted", {})
    ph_ord, named = phase_order(tasks)
    last = named[-1] if named else None
    gate_last = cfg.get("gate_last_phase", False) and last

    @lru_cache(maxsize=None)
    def down(tid):
        return tasks[tid]["days"] + max((down(k) for k in kids[tid]), default=0)

    for t in tasks.values():
        if t["role0"] not in alloc:
            alloc[t["role0"]] = 1.0  # unknown owner treated as full-time, and reported

    free, fin, placed = defaultdict(float), {}, set()
    prev_on = defaultdict(lambda: None)
    todo = sorted(tasks.values(), key=lambda t: (ph_ord[t["phase"]], -down(t["id"])))
    for _ in range(len(tasks) + 2):
        if len(placed) == len(tasks):
            break
        moved = False
        for t in todo:
            if t["id"] in placed or any(d not in placed for d in t["deps"]):
                continue
            est = max((fin[d] for d in t["deps"]), default=0.0)
            if gate_last and t["phase"] == last:
                prior = [
                    fin[x] for x in placed if ph_ord[tasks[x]["phase"]] < ph_ord[last]
                ]
                est = max([est] + ([max(prior)] if prior else []))
            pooled = level and roles.get(t["role0"], {}).get("pool", False)
            if not pooled:
                cand = [t["role0"]]
            elif t["ws"] in restricted:
                cand = [r for r in restricted[t["ws"]] if r in alloc] or [t["role0"]]
            else:
                cand = pool or [t["role0"]]
            role = min(cand, key=lambda r: max(est, free[r]) + t["days"] / alloc[r])
            start = max(est, free[role])
            end = start + t["days"] / alloc[role]
            # Record what actually held this task back: a dependency, or a busy seat. In a
            # resource-constrained schedule the binding chain is usually a mix of both, so
            # following dependency edges alone understates it badly.
            cause, why = None, "start"
            if t["deps"]:
                lead = max(t["deps"], key=lambda d: fin[d])
                if abs(fin[lead] - start) < 1e-9:
                    cause, why = lead, "dependency"
            if cause is None and prev_on[role] and abs(free[role] - start) < 1e-9:
                cause, why = prev_on[role], "waiting for seat"
            t["cause"], t["why"] = cause, why
            prev_on[role] = t["id"]
            free[role] = fin[t["id"]] = end
            t.update(role=role, start=start, end=end)
            placed.add(t["id"])
            moved = True
        if not moved:
            sys.exit(
                "error: schedule stuck — unresolved: "
                + ", ".join(sorted(set(tasks) - placed))[:200]
            )
    return fin, ph_ord, named, alloc


def binding_chain(tasks, fin):
    """The chain that actually sets the finish date: dependency waits AND seat waits."""
    node = max(fin, key=lambda t: fin[t])
    chain, guard = [node], 0
    while tasks[node].get("cause") and guard < len(tasks) + 2:
        node = tasks[node]["cause"]
        chain.append(node)
        guard += 1
    return list(reversed(chain))


def report(tasks, fin, ph_ord, named, alloc, cfg, label=""):
    weeks = max(fin.values()) / 5
    total = sum(t["days"] for t in tasks.values())
    print(f"\n{'=' * 74}")
    print(
        f"{label or 'SCHEDULE'}: {len(tasks)} tasks · {total:g} days · {weeks:.1f} weeks"
    )
    print("=" * 74)
    busy = defaultdict(float)
    for t in tasks.values():
        busy[t["role"]] += t["days"]
    print(f"\n{'Seat':28}{'alloc':>7}{'task d':>9}{'booked d':>10}{'util':>7}")
    for r in sorted(busy, key=lambda x: -busy[x]):
        cap = alloc[r] * weeks * 5
        flag = "" if r in cfg["roles"] else "  <- not in team.json"
        print(
            f"{r:28}{alloc[r]:>7.2f}{busy[r]:>9.1f}{cap:>10.1f}"
            f"{busy[r] / cap * 100:>6.0f}%{flag}"
        )
    print(f"\n{'Phase':12}{'days':>8}{'tasks':>7}{'starts wk':>11}{'ends wk':>9}")
    for p in named:
        ts = [t for t in tasks.values() if t["phase"] == p]
        if ts:
            print(
                f"{p:12}{sum(t['days'] for t in ts):>8g}{len(ts):>7}"
                f"{min(t['start'] for t in ts) / 5:>11.1f}"
                f"{max(t['end'] for t in ts) / 5:>9.1f}"
            )
    loe = cfg.get("loe_days_per_week", {})
    if loe:
        print("\nLevel-of-effort roles (outside the backlog):")
        for r, dpw in loe.items():
            print(
                f"  {r:28}{dpw} d/wk × {weeks:.1f} wks = {dpw * weeks:.0f} person-days"
            )
        extra = sum(d * weeks for d in loe.values())
        print(
            f"  total delivery cost = {total:g} + {extra:.0f} = {total + extra:.0f} person-days"
        )
    cp = binding_chain(tasks, fin)
    dep_w = sum(1 for t in cp if tasks[t]["why"] == "dependency")
    seat_w = sum(1 for t in cp if tasks[t]["why"] == "waiting for seat")
    print(
        f"\nBinding chain — what sets the {weeks:.1f}-week finish ({len(cp)} tasks, "
        f"{sum(tasks[t]['days'] for t in cp):g} days; {dep_w} dependency waits, "
        f"{seat_w} seat waits):"
    )
    NOTE = {
        "dependency": "after {}",
        "waiting for seat": "seat busy with {}",
        "start": "starts immediately",
    }
    for t in cp:
        tk = tasks[t]
        note = NOTE[tk["why"]].format(tk["cause"])
        print(
            f"    wk {tk['start'] / 5:5.1f} → {tk['end'] / 5:5.1f}  {t:7}"
            f"{tk['days']:>5g}d  {tk['role']:24} ({note})"
        )
    if seat_w > dep_w:
        print(
            "  NOTE: more of this chain is waiting for PEOPLE than waiting for WORK. "
            "Re-levelling or raising an allocation will shorten the plan more than cutting scope."
        )
    part = [r for r in busy if alloc[r] < 1.0]
    if part:
        print("\nPart-time seats multiply calendar time:")
        for r in sorted(part, key=lambda x: alloc[x]):
            print(
                f"  {r:28}{alloc[r]:.0%} → 1 task-day takes {1 / alloc[r]:.0f} calendar days"
                f"; {busy[r]:g}d spans {busy[r] / (alloc[r] * 5):.1f} wks"
            )
    return weeks


def write_gantt(wb, name, tasks, weeks, after=None):
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter

    DARK = PatternFill("solid", fgColor="FF171725")
    LIGHT = PatternFill("solid", fgColor="FFF7F7F9")
    THIN = Border(bottom=Side(style="thin", color="FFE0E0E4"))
    BARS = ["FF4F81BD", "FF6AA84F", "FFF79646", "FFC00000", "FF8064A2", "FF4BACC6"]
    _, named = phase_order(tasks)
    colour = {
        p: PatternFill("solid", fgColor=BARS[i % len(BARS)])
        for i, p in enumerate(named)
    }
    if name in wb.sheetnames:
        del wb[name]
    idx = (
        wb.sheetnames.index(after) + 1 if after in wb.sheetnames else len(wb.sheetnames)
    )
    g = wb.create_sheet(name, idx)
    g["A1"] = "GANTT — RESOURCE-CONSTRAINED SCHEDULE"
    g["A1"].font = Font(name="Calibri", sz=14, bold=True, color="FF171725")
    c = g.cell(row=2, column=1)
    c.value = (
        f"{weeks:.1f} weeks. One person, one task at a time, at their allocation — a 50% seat "
        f"needs 2 calendar days per task-day. A task starts when its dependencies are done and "
        f"its owner is free. Bar colour = phase. Generated from the task backlog; every date in "
        f"the workbook must come from this sheet."
    )
    c.font = Font(name="Calibri", sz=10, color="FF525259")
    c.alignment = Alignment(wrap_text=True, vertical="top")
    FIX = ["Phase", "ID", "Task", "Owner", "Days", "Start wk", "End wk"]
    nw = int(weeks) + (1 if weeks % 1 else 0)
    for i, lab in enumerate(FIX + [f"W{w + 1}" for w in range(nw)]):
        cc = g.cell(row=4, column=i + 1)
        cc.value, cc.fill = lab, DARK
        cc.font = Font(
            name="Calibri", sz=8 if i >= len(FIX) else 9, bold=True, color="FFFFFFFF"
        )
        cc.alignment = Alignment(
            wrap_text=True, vertical="bottom", horizontal="center" if i >= 4 else "left"
        )
    r = 5
    for t in sorted(tasks.values(), key=lambda x: (x["start"], x["id"])):
        vals = [
            t["phase"],
            t["id"],
            t["task"],
            t["role"],
            t["days"],
            round(t["start"] / 5, 1),
            round(t["end"] / 5, 1),
        ]
        for i, v in enumerate(vals):
            cc = g.cell(row=r, column=i + 1)
            cc.value = v
            cc.font = Font(name="Calibri", sz=9, color="FF333333")
            cc.alignment = Alignment(
                wrap_text=True,
                vertical="top",
                horizontal="center" if i >= 4 else "left",
            )
            cc.border = THIN
            if r % 2 == 0:
                cc.fill = LIGHT
        sw, ew = t["start"] / 5, t["end"] / 5
        for w in range(nw):
            if sw < w + 1 and ew > w:
                cc = g.cell(row=r, column=len(FIX) + 1 + w)
                cc.fill = colour.get(t["phase"], LIGHT)
                cc.border = THIN
        g.row_dimensions[r].height = None
        r += 1
    r += 1
    g.cell(row=r, column=1).value = "LEGEND"
    g.cell(row=r, column=1).font = Font(
        name="Calibri", sz=10, bold=True, color="FFF01446"
    )
    for i, p in enumerate(named):
        cc = g.cell(row=r, column=3 + i)
        cc.value, cc.fill = p, colour[p]
        cc.font = Font(name="Calibri", sz=9, bold=True, color="FFFFFFFF")
        cc.alignment = Alignment(horizontal="center")
    for col, w in zip("ABCDEFG", [10, 8, 48, 22, 7, 9, 8]):
        g.column_dimensions[col].width = w
    for w in range(nw):
        g.column_dimensions[get_column_letter(len(FIX) + 1 + w)].width = 3.2
    g.freeze_panes = "H5"
    return g


def main():
    ap = argparse.ArgumentParser(
        description=__doc__, formatter_class=argparse.RawDescriptionHelpFormatter
    )
    ap.add_argument("--backlog", required=True)
    ap.add_argument("--sheet", default="02 Task Backlog")
    ap.add_argument("--alloc", required=True, help="team.json")
    ap.add_argument(
        "--level",
        action="store_true",
        help="send each task to the seat that finishes it soonest",
    )
    ap.add_argument("--scenarios", action="store_true", help="try allocation changes")
    ap.add_argument(
        "--gantt", metavar="SHEET", help="write a Gantt sheet with this name"
    )
    ap.add_argument(
        "--gantt-after", metavar="SHEET", help="place the Gantt after this sheet"
    )
    ap.add_argument(
        "--apply", action="store_true", help="write to the file (default is dry run)"
    )
    a = ap.parse_args()

    cfg = json.loads(Path(a.alloc).read_text())
    wb, sheet, ws, hdr_row, M, tasks = load(a.backlog, a.sheet)
    kids, problems = wire(tasks)
    if problems:
        print("DEPENDENCY GRAPH PROBLEMS — fix these before trusting any date:")
        for p in problems:
            print("  ✗ " + p)
        sys.exit(1)
    print(f"loaded {len(tasks)} tasks from {sheet!r} (header row {hdr_row})")

    fin, ph_ord, named, alloc = run(tasks, kids, cfg, level=a.level)
    weeks = report(
        tasks,
        fin,
        ph_ord,
        named,
        alloc,
        cfg,
        "LEVELLED SCHEDULE" if a.level else "SCHEDULE AS OWNED",
    )

    if a.scenarios:
        print(
            f"\n{'=' * 74}\nSCENARIOS — measure the lever, do not guess it\n{'=' * 74}"
        )
        base = weeks
        print(f"  {'as configured':52}{base:6.2f} wks")
        import copy

        for label, mut in [
            ("level work across the pooled seats", lambda c: c),
            *[
                (
                    f"{r} at 100%",
                    (lambda rr: lambda c: c["roles"][rr].update(alloc=1.0))(r),
                )
                for r, v in cfg["roles"].items()
                if v["alloc"] < 1.0
            ],
            (
                "every part-time seat at 100%",
                lambda c: [
                    v.update(alloc=1.0) for v in c["roles"].values() if v["alloc"] < 1.0
                ],
            ),
        ]:
            c2 = copy.deepcopy(cfg)
            mut(c2)
            t2 = {k: dict(v) for k, v in tasks.items()}
            for t in t2.values():
                t["role0"] = tasks[t["id"]]["role0"]
            k2, _ = wire(t2)
            f2, *_ = run(t2, k2, c2, level=True)
            w2 = max(f2.values()) / 5
            delta = base - w2
            print(f"  {label:52}{w2:6.2f} wks  {delta:+.2f}")

    if a.gantt or (a.level and a.apply):
        if not a.apply:
            print("\n[dry run — pass --apply to write]")
            return
        bak = Path(a.backlog).with_name(f".bak-{Path(a.backlog).name}")
        shutil.copy2(a.backlog, bak)
        print(f"\nbacked up -> {bak.name}")
        if a.level and "role" in M:
            for i, tid in enumerate(
                [
                    str(ws.cell(row=r, column=M["id"] + 1).value)
                    for r in range(hdr_row + 1, ws.max_row + 1)
                    if ws.cell(row=r, column=M["id"] + 1).value
                ]
            ):
                if tid in tasks:
                    ws.cell(row=hdr_row + 1 + i, column=M["role"] + 1).value = tasks[
                        tid
                    ]["role"]
            print("wrote levelled Owner back into the backlog")
        if a.gantt:
            write_gantt(wb, a.gantt, tasks, weeks, a.gantt_after)
            print(f"wrote {a.gantt!r}")
        wb.calculation.fullCalcOnLoad = True
        wb.save(a.backlog)
        print("saved")
    elif not a.scenarios:
        print("\n[nothing written — add --gantt NAME --apply to write the Gantt]")


if __name__ == "__main__":
    main()
